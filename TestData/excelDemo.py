import openpyxl
book = openpyxl.load_workbook("C:\\Users\\Admin\\Desktop\\excelDemo.xlsx")
data = {} #empty dictionery
sheet = book.active
i= sheet.max_row
j= sheet.max_column
for i in range (1, i+1):

    if(sheet.cell(row=i , column= 1).value == "Testcase1"):
        for j in range(2, j+1):
            print(f"{sheet.cell(row=1 , column=j).value}  : {sheet.cell(row = i , column= j ).value}")
            data[sheet.cell(row=1 , column=j).value] = sheet.cell(row=i , column=j).value

print(data)















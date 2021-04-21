import openpyxl
class TestPageData:


    test_data = [{"name" :"Harsh", "email":"harsh@gmail.com" ,"gender": "Male"} , {"name": "Shivani" , "email" :"shivani@gmail.com" , "gender":"Female"}]
    @staticmethod    #call static method without creating object.
    def gettestdata(test_data):
        book = openpyxl.load_workbook("C:\\Users\\Admin\\Desktop\\excelDemo.xlsx")
        data = {}  # empty dictionery
        sheet = book.active
        i = sheet.max_row
        j = sheet.max_column
        for i in range(1, i + 1):

            if (sheet.cell(row=i, column=1).value == test_data):
                for j in range(2, j + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [data]


